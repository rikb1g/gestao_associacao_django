from decimal import Decimal
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.core.files import File
from django.core.exceptions import ObjectDoesNotExist
from .models import Salarios
from apps.funcionarios.models import Funcionario
from apps.documentos.models import RecibosVencimento
from apps.gestao.models import Despesas,TipoDespesa
from openpyxl import load_workbook
from datetime import date

"""
@receiver(post_save, sender=Salarios)
def criar_recibo(sender, instance, created, **kwargs):
    if created:
        guardar_recibo(instance.funcionario.nome, instance.
        data_inicio, instance.data_fim, instance.valor)
        caminho_do_arquivo = f'media/recibos/{instance.funcionario.nome}_{instance.data_fim}.xlsx'

        funcionario = Funcionario.objects.get(pk= instance.funcionario.pk)
        novo_recibo = RecibosVencimento.objects.create(
            funcionario = funcionario,
            data = instance.data_fim,
            ficheiro = File(open(caminho_do_arquivo,'rb'))
        )
        
"""

@receiver(post_save,sender=Salarios)
def calcular_salario(sender,instance, created, **kwargs):
    if created:
        print(instance.valor)
        if instance.funcionario.duodecimos :
            duodecimos = (instance.funcionario.salario / 12) * 2
            salario_sem_descontos = duodecimos + instance.valor
            seg_social = salario_sem_descontos * Decimal("0.11")
            salario_final = salario_sem_descontos - seg_social
            print(salario_final)
            escola= instance.funcionario.escola
            data = date.today()
            print(data)
            try:
                tipo_despesa = TipoDespesa.objects.get(nome="Ordenado")
                print(tipo_despesa)
            
                processar_despesa = Despesas(tipo= tipo_despesa,valor = salario_final,escola = escola,data = data,descricao= f"Salario {instance.funcionario.nome}")
                processar_despesa.save()
                print("Encontrou tipo despesa")
            
            except ObjectDoesNotExist:
                novo_tipo = TipoDespesa(nome = "Ordenado",escola = escola)
                novo_tipo.save()
                print("Nao encontrou tipo despesa")

                processar_despesa = Despesas(tipo = novo_tipo,valor = salario_final,escola = escola, data = data,descricao= f"Salario {instance.funcionario.nome}")
                processar_despesa.save()
                print("Gravou com sucesso")

        else:

            print("Sem duodecimos")
            salario_iliquido = instance.valor
            print(salario_iliquido)
            seg_social = salario_iliquido * Decimal('0.11')
            print(seg_social)
            
            salario_final = salario_iliquido - seg_social
            salario_final = round(salario_final, 2)
            escola = instance.funcionario.escola
            data = date.today()
            
            try:
                tipo_despesa = TipoDespesa.objects.get(nome="Ordenado")
                processar_despesa = Despesas(tipo= tipo_despesa,valor = salario_final,descricao= f"Salario {instance.funcionario.nome}",escola = escola, data = data)
                processar_despesa.save()
                print("Encontrou tipo despesa")
            except ObjectDoesNotExist:
                novo_tipo = TipoDespesa(nome = "Ordenado",escola = escola)
                novo_tipo.save()
                print("Nao encontrou tipo despesa")

                processar_despesa = Despesas(tipo = novo_tipo,valor = salario_final,escola = escola,data = data,descricao= f"Salario {instance.funcionario.nome}")
                processar_despesa.save()
                print("Gravou com sucesso")
                
            
            











        





def guardar_recibo(nome, data_inicio, data_fim, valor):
    try:
        wb = load_workbook("media/recibos/exemplorecibo.xlsx")
        ws = wb.active
    

        ws['B7'] = nome
        arquivo_salvo = f'media/recibos/{nome}_{data_fim}.xlsx'
        wb.save(arquivo_salvo)
        print("Arquivo salvo com sucesso!")
        
    except FileNotFoundError:
        print("Arquivo exemplorecibo.xlsx não encontrado")
    except Exception as error:
        print(f"Ocorreu um erro {error}")
       
        
